#!/usr/bin/env python3
"""
Generate data.json for the BizX Content Explorer dashboard.

Usage:
    python scripts/generate_data.py [path/to/analytics.xlsx]

If no path given, looks for data/analytics.xlsx.
Reads navigator_analysis.json from data/ and outputs src/data.json.
"""
import json, os, re, sys

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl..."); os.system(f"{sys.executable} -m pip install openpyxl -q"); import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = sys.argv[1] if len(sys.argv) > 1 else os.path.join(ROOT, "data", "analytics.xlsx")
NAV = os.path.join(ROOT, "data", "navigator_analysis.json")
OUT = os.path.join(ROOT, "src", "data.json")

def read_sheet(wb, name):
    if name not in wb.sheetnames: return []
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows

print(f"Reading XLSX: {XLSX}")
wb = openpyxl.load_workbook(XLSX, data_only=True)

print(f"Reading navigator analysis: {NAV}")
nav = json.load(open(NAV))

# ============================================================
# Parse XLSX sheets
# ============================================================
xlsx_industries = read_sheet(wb, "Industries")
xlsx_sectors = read_sheet(wb, "Sectors")
xlsx_phases = read_sheet(wb, "Operating Phases")
xlsx_tasks = read_sheet(wb, "Task Progress Metrics")
xlsx_neq = read_sheet(wb, "Non Essential Questions")
xlsx_legal = read_sheet(wb, "Legal Structures")

# Build user counts by industry name -> count
# XLSX names are multi-line: "Industry Name\nDescription". Take first line only.
industry_users = {}
unknown_industry = 0
unmatched_industries = []
for row in xlsx_industries:
    raw = (row.get("Industry") or "").strip()
    if not raw: continue
    name = raw.split("\n")[0].strip()
    cnt = int(row.get("Count") or 0)
    if name == "Unknown":
        unknown_industry = cnt
        continue
    industry_users[name] = cnt

# Build sector user counts
sector_users = {}
unknown_sector = 0
for row in xlsx_sectors:
    raw = (row.get("Sector") or "").strip()
    if not raw: continue
    name = raw.split("\n")[0].strip()
    cnt = int(row.get("Count") or 0)
    if name == "Unknown":
        unknown_sector = cnt
        continue
    sector_users[name] = cnt

# Operating phases
phases = {}
total_businesses = 0
unknown_phase = 0
for row in xlsx_phases:
    phase = (row.get("Phase") or "").strip()
    if not phase: continue
    cnt = int(row.get("Count") or 0)
    if phase == "Unknown":
        unknown_phase = cnt
        continue
    phases[phase] = cnt
    total_businesses += cnt

# Task progress
task_progress = []
for row in xlsx_tasks:
    task = row.get("Task") or ""
    task_progress.append({
        "task": task.strip(),
        "completed": int(row.get("COMPLETED") or 0),
        "todo": int(row.get("TO_DO") or 0),
        "last45": int(row.get("Completed Last 45 Days") or 0),
        "avgTime": (row.get("Avg Time to Complete") or ""),
        "total": int(row.get("Total") or 0),
    })
task_progress.sort(key=lambda x: -x["total"])

# Build task name -> slug mapping from navigator analysis
task_name_to_slug = nav.get("taskNameToSlug", {})

# ============================================================
# CODEBASE-FIRST: Build master task inventory from markdown files
# ============================================================
# The codebase is the source of truth for what tasks exist.
# XLSX analytics are overlaid where available.
nav_task_content_root = os.path.join(ROOT, "..", "nav", "content", "src", "roadmaps")
_task_dirs = [
    os.path.join(nav_task_content_root, "tasks"),
    os.path.join(nav_task_content_root, "license-tasks"),
    os.path.join(nav_task_content_root, "env-tasks"),
]

codebase_task_inventory = {}  # task_id -> {name, urlSlug, dir}
for _td in _task_dirs:
    if not os.path.isdir(_td): continue
    import glob as _g
    for _f in _g.glob(os.path.join(_td, "*.md")):
        _content = open(_f).read()
        _parts = _content.split("---")
        if len(_parts) < 3: continue
        _fm = _parts[1]
        _tid = _tname = _turl = None
        for _line in _fm.split("\n"):
            _line = _line.strip()
            if _line.startswith("id:"): _tid = _line[3:].strip().strip('"').strip("'")
            elif _line.startswith("name:"): _tname = _line[5:].strip().strip('"').strip("'")
            elif _line.startswith("urlSlug:"): _turl = _line[8:].strip().strip('"').strip("'")
        if _tid:
            codebase_task_inventory[_tid] = {
                "name": _tname or _tid,
                "urlSlug": _turl or _tid,
                "dir": os.path.basename(_td),
            }

if codebase_task_inventory:
    print(f"  Codebase task inventory: {len(codebase_task_inventory)} markdown files")

# Count fundings and certifications from codebase
_funding_dir = os.path.join(ROOT, "..", "nav", "content", "src", "fundings")
_cert_dir = os.path.join(ROOT, "..", "nav", "content", "src", "certifications")
import glob as _g3
codebase_funding_count = len(_g3.glob(os.path.join(_funding_dir, "*.md"))) if os.path.isdir(_funding_dir) else 0
codebase_cert_count = len(_g3.glob(os.path.join(_cert_dir, "*.md"))) if os.path.isdir(_cert_dir) else 0
if codebase_funding_count:
    print(f"  Codebase fundings: {codebase_funding_count}, certifications: {codebase_cert_count}")

    # Build lookup: XLSX slug -> XLSX row(s)
    xlsx_by_slug = {}  # slug -> list of XLSX rows
    for tp_entry in task_progress:
        slug = task_name_to_slug.get(tp_entry["task"], tp_entry["task"])
        if slug not in xlsx_by_slug:
            xlsx_by_slug[slug] = []
        xlsx_by_slug[slug].append(tp_entry)

    # Also build by display name -> list of possible task IDs from codebase
    name_to_codebase_ids = {}
    for tid, info in codebase_task_inventory.items():
        n = info["name"]
        if n not in name_to_codebase_ids:
            name_to_codebase_ids[n] = []
        name_to_codebase_ids[n].append(tid)

    # Build merged task_progress: start from codebase, overlay XLSX
    merged_tasks = []
    matched_xlsx_slugs = set()

    for tid, info in codebase_task_inventory.items():
        # Try to find matching XLSX data
        xlsx_match = None
        # 1. Direct slug match
        if tid in xlsx_by_slug:
            xlsx_match = xlsx_by_slug[tid]
            matched_xlsx_slugs.add(tid)
        else:
            # 2. Try matching via display name
            for tp_entry in task_progress:
                entry_slug = task_name_to_slug.get(tp_entry["task"], tp_entry["task"])
                if entry_slug == tid:
                    xlsx_match = [tp_entry]
                    matched_xlsx_slugs.add(entry_slug)
                    break
            # 3. Try matching by name -> all possible IDs
            if not xlsx_match:
                for tp_entry in task_progress:
                    possible_ids = name_to_codebase_ids.get(tp_entry["task"], [])
                    if tid in possible_ids:
                        entry_slug = task_name_to_slug.get(tp_entry["task"], tp_entry["task"])
                        if entry_slug not in matched_xlsx_slugs:
                            xlsx_match = [tp_entry]
                            matched_xlsx_slugs.add(entry_slug)
                            break

        if xlsx_match:
            # Use the XLSX row with the most data
            best = max(xlsx_match, key=lambda x: x["total"])
            merged_tasks.append({
                "task": info["name"],
                "slug": tid,
                "urlSlug": info["urlSlug"],
                "completed": best["completed"],
                "todo": best["todo"],
                "last45": best["last45"],
                "avgTime": best["avgTime"],
                "total": best["total"],
                "hasAnalytics": True,
            })
        else:
            # Codebase task with no XLSX data
            merged_tasks.append({
                "task": info["name"],
                "slug": tid,
                "urlSlug": info["urlSlug"],
                "completed": 0,
                "todo": 0,
                "last45": 0,
                "avgTime": "",
                "total": 0,
                "hasAnalytics": False,
            })

    # Add XLSX-only entries (not in codebase) — these are retired/legacy
    for tp_entry in task_progress:
        slug = task_name_to_slug.get(tp_entry["task"], tp_entry["task"])
        if slug not in matched_xlsx_slugs and slug not in codebase_task_inventory:
            merged_tasks.append({
                "task": tp_entry["task"],
                "slug": slug,
                "urlSlug": slug,
                "completed": tp_entry["completed"],
                "todo": tp_entry["todo"],
                "last45": tp_entry["last45"],
                "avgTime": tp_entry["avgTime"],
                "total": tp_entry["total"],
                "hasAnalytics": True,
            })

    merged_tasks.sort(key=lambda x: -x["total"])
    no_analytics = sum(1 for t in merged_tasks if not t["hasAnalytics"])
    xlsx_only = sum(1 for t in merged_tasks if t["hasAnalytics"] and t["slug"] not in codebase_task_inventory)
    print(f"  Merged task list: {len(merged_tasks)} total ({len(merged_tasks)-no_analytics-xlsx_only} matched, {no_analytics} codebase-only, {xlsx_only} XLSX-only/legacy)")

    # Replace task_progress with the merged list
    task_progress = merged_tasks

# Non-essential questions
neq_data = []
for row in xlsx_neq:
    q = (row.get("Question Name") or "").strip()
    if not q: continue
    neq_data.append({
        "question": q,
        "yes": int(row.get("Yes Count") or 0),
        "no": int(row.get("No Count") or 0),
        "unknown": int(row.get("Unknown Count") or 0),
    })

# Legal structures
legal_data = []
unknown_legal = 0
for row in xlsx_legal:
    cat = (row.get("Category") or "").strip()
    if not cat: continue
    cnt = int(row.get("Count") or 0)
    if cat == "Unknown":
        unknown_legal = cnt
        continue
    legal_data.append({"structure": cat, "count": cnt})

# ============================================================
# Compute content reachability from navigator analysis
# ============================================================
nav_industries = nav["industries"]
nav_sectors = nav["sectors"]
nav_aa = nav["anytimeActions"]
nav_fundings = nav["fundings"]
nav_mismatches = nav["sectorMismatches"]

# Patch nav_industries with live sector data from industry JSON files if available
_live_ind_dir = os.path.join(ROOT, "..", "nav", "content", "src", "roadmaps", "industries")
if os.path.isdir(_live_ind_dir):
    _sector_updates = 0
    _mismatch_updates = []
    import glob as _g2
    for _f in _g2.glob(os.path.join(_live_ind_dir, "*.json")):
        _idata = json.load(open(_f))
        _iid = _idata.get("id")
        if not _iid or not _idata.get("isEnabled"): continue
        if _iid in nav_industries:
            old_sector = nav_industries[_iid].get("defaultSectorId", "")
            new_sector = _idata.get("defaultSectorId", "")
            if old_sector != new_sector and new_sector:
                nav_industries[_iid]["defaultSectorId"] = new_sector
                _sector_updates += 1
            # Also refresh other fields that might have changed
            nav_industries[_iid]["nonEssentialQuestionsIds"] = _idata.get("nonEssentialQuestionsIds", [])
            nav_industries[_iid]["industryOnboardingQuestions"] = _idata.get("industryOnboardingQuestions", {})
    if _sector_updates:
        print(f"  Patched {_sector_updates} industry sector assignments from live codebase")
    # Recompute mismatches from original nav data (sectors may have changed)
    _original_mismatches = nav["sectorMismatches"]  # dict: ind_id -> suggested_sector
    # Intentional sector assignments (confirmed by team, not mismatches)
    _intentional = {"petcare"}  # Pet Care intentionally assigned to other-services
    _new_mismatches = {}
    for _iid, _ind in nav_industries.items():
        if not _ind.get("isEnabled"): continue
        if _iid in _intentional: continue
        if _iid in _original_mismatches:
            assigned = _ind.get("defaultSectorId", "")
            suggested = _original_mismatches[_iid]
            if assigned != suggested:
                _new_mismatches[_iid] = suggested
    nav_mismatches = _new_mismatches
    if len(nav_mismatches) != len(_original_mismatches):
        print(f"  Sector mismatches updated: {len(_original_mismatches)} → {len(nav_mismatches)} (sectors corrected in codebase)")

reachable_sectors = set(i["defaultSectorId"] for i in nav_industries.values() if i.get("isEnabled"))
orphaned_sectors = set(nav_sectors.keys()) - reachable_sectors

def count_aa(industry_id, sector_id):
    total, by_ind, by_sec, by_uni = 0, 0, 0, 0
    names_ind, names_sec, names_uni = [], [], []
    for aa in nav_aa:
        if aa.get("applyToAllUsers"):
            total += 1; by_uni += 1; names_uni.append(aa["name"]); continue
        iids = aa.get("industryIds", [])
        sids = aa.get("sectorIds", [])
        if industry_id and iids and industry_id in iids:
            total += 1; by_ind += 1; names_ind.append(aa["name"])
        elif sector_id and sids and sector_id in sids:
            total += 1; by_sec += 1; names_sec.append(aa["name"])
    return total, by_ind, by_sec, by_uni, names_ind, names_sec, names_uni

def count_fundings(sector_id):
    names = []
    for f in nav_fundings:
        secs = f.get("sector", [])
        if not secs:
            names.append(f["name"])
        elif sector_id and any(re.match(s, sector_id, re.I) for s in secs):
            names.append(f["name"])
    return len(names), names

# ============================================================
# Compute roadmap task differentiation
# ============================================================
universal_tasks = nav.get("universalTasks", [])

# Build task -> industry mapping (excluding universal tasks)
task_to_industries = {}
enabled_industries = {k: v for k, v in nav_industries.items() if v.get("isEnabled")}
for ind_id, ind in enabled_industries.items():
    for task in ind.get("roadmapTaskNames", []):
        if task in universal_tasks: continue
        if task not in task_to_industries:
            task_to_industries[task] = []
        task_to_industries[task].append(ind_id)

# For each industry, classify tasks as unique (only this industry) or shared
industry_task_analysis = {}
for ind_id, ind in enabled_industries.items():
    diff_tasks = [t for t in ind.get("roadmapTaskNames", []) if t not in universal_tasks]
    unique = [t for t in diff_tasks if len(task_to_industries.get(t, [])) == 1]
    shared = [t for t in diff_tasks if len(task_to_industries.get(t, [])) > 1]
    industry_task_analysis[ind_id] = {
        "allTasks": diff_tasks,
        "uniqueTaskNames": unique,
        "sharedTaskNames": shared,
        "totalTasks": len(diff_tasks),
        "uniqueTasks": len(unique),
        "sharedTasks": len(shared),
    }

# Build task frequency for output
task_frequency = {}
for task, inds_list in sorted(task_to_industries.items()):
    task_frequency[task] = {"count": len(inds_list), "industries": inds_list}

# Categorize each task in task_progress: universal, shared, unique, addon, or uncategorized
# Add-on tasks get colored by their potential reach (how many industries could trigger them)
addon_slugs = set(nav.get("addonTaskSlugs", []))
addon_reach = nav.get("addonTaskReach", {})
total_enabled = len([i for i in nav_industries.values() if i.get("isEnabled")])
for tp_entry in task_progress:
    name = tp_entry["task"]
    slug = tp_entry.get("slug") or task_name_to_slug.get(name, name)
    tp_entry["slug"] = slug  # ensure it's always there
    is_addon = slug in addon_slugs
    reach = addon_reach.get(slug, 0) if is_addon else 0
    if slug in universal_tasks or name in universal_tasks:
        tp_entry["category"] = "universal"
        tp_entry["isAddon"] = False
        tp_entry["reachCount"] = total_enabled
    elif slug in task_frequency:
        tp_entry["category"] = "shared" if task_frequency[slug]["count"] > 1 else "unique"
        tp_entry["isAddon"] = False
        tp_entry["reachCount"] = task_frequency[slug]["count"]
    elif name in task_frequency:
        tp_entry["category"] = "shared" if task_frequency[name]["count"] > 1 else "unique"
        tp_entry["isAddon"] = False
        tp_entry["reachCount"] = task_frequency[name]["count"]
    elif is_addon:
        if reach >= total_enabled:
            tp_entry["category"] = "universal"
        elif reach > 1:
            tp_entry["category"] = "shared"
        else:
            tp_entry["category"] = "unique"
        tp_entry["isAddon"] = True
        tp_entry["reachCount"] = reach
    else:
        tp_entry["category"] = "uncategorized"
        tp_entry["isAddon"] = False
        tp_entry["reachCount"] = 0

cat_counts = {}
for tp_entry in task_progress:
    c = tp_entry["category"]
    cat_counts[c] = cat_counts.get(c, 0) + 1
print(f"  Task categories: {cat_counts}")

# ============================================================
# Build combined industry data
# ============================================================
# Distinct paths per industry (from navigator analysis)
distinct_paths_map = {}
for dp in nav.get("distinctPaths", []):
    distinct_paths_map[dp["id"]] = dp["distinctPaths"]
total_distinct_paths = nav.get("totalDistinctPaths", 0)

combined_industries = []
for ind_id, ind in sorted(nav_industries.items(), key=lambda x: x[1]["name"]):
    if not ind.get("isEnabled"): continue

    assigned_sector = ind["defaultSectorId"]
    assigned_sector_name = nav_sectors.get(assigned_sector, {}).get("name", assigned_sector)
    suggested_sector = nav_mismatches.get(ind_id)
    suggested_sector_name = nav_sectors.get(suggested_sector, {}).get("name", "") if suggested_sector else ""

    # Match user count from XLSX by industry name (strip trailing spaces)
    users = industry_users.get(ind["name"].strip(), 0)

    # AA counts for STARTING path (has industryId + assigned sector)
    s_total, s_ind, s_sec, s_uni, s_ni, s_ns, s_nu = count_aa(ind_id, assigned_sector)
    s_fund, s_fnames = count_fundings(assigned_sector)

    # AA counts for OWNING path (no industryId, uses likely correct sector)
    owning_sector = suggested_sector or assigned_sector
    o_total, o_ind, o_sec, o_uni, o_ni, o_ns, o_nu = count_aa(None, owning_sector)
    o_fund, o_fnames = count_fundings(owning_sector)

    # What would STARTING see if sector were fixed?
    if suggested_sector:
        f_total, f_ind, f_sec, f_uni, f_ni, f_ns, f_nu = count_aa(ind_id, suggested_sector)
        f_fund, f_fnames = count_fundings(suggested_sector)
        missed_aa = f_total - s_total
        missed_fund = f_fund - s_fund
    else:
        f_total, missed_aa, missed_fund = s_total, 0, 0

    ta = industry_task_analysis.get(ind_id, {})

    combined_industries.append({
        "id": ind_id,
        "name": ind["name"],
        "users": users,
        "sector": assigned_sector,
        "sectorName": assigned_sector_name,
        "sectorMismatch": suggested_sector or None,
        "sectorMismatchName": suggested_sector_name or None,
        "roadmapTasks": ind["roadmapStepCount"],
        "nonEssentialQs": len(ind.get("nonEssentialQuestionsIds", [])),
        "distinctPaths": distinct_paths_map.get(ind_id, 0),
        "totalDiffTasks": ta.get("totalTasks", 0),
        "uniqueTasks": ta.get("uniqueTasks", 0),
        "sharedTasks": ta.get("sharedTasks", 0),
        "uniqueTaskNames": ta.get("uniqueTaskNames", []),
        "sharedTaskNames": ta.get("sharedTaskNames", []),
        "allDiffTasks": ta.get("allTasks", []),
        "starting": {
            "aaTotal": s_total, "aaByIndustry": s_ind, "aaBySector": s_sec, "aaUniversal": s_uni,
            "aaNamesByIndustry": s_ni, "aaNamesBySector": s_ns,
            "fundings": s_fund,
        },
        "owning": {
            "sector": owning_sector,
            "sectorName": nav_sectors.get(owning_sector,{}).get("name", owning_sector),
            "aaTotal": o_total, "aaBySector": o_sec, "aaUniversal": o_uni,
            "aaNamesBySector": o_ns,
            "fundings": o_fund,
        },
        "ifFixed": {
            "aaTotal": f_total if suggested_sector else None,
            "fundings": f_fund if suggested_sector else None,
            "missedAA": missed_aa,
            "missedFund": missed_fund,
        },
    })

# ============================================================
# Build sector summary
# ============================================================
sector_summary = []
for sid, sdata in sorted(nav_sectors.items(), key=lambda x: x[1]["name"]):
    mapped_industries = [i["id"] for i in combined_industries if i["sector"] == sid]
    mapped_users = sum(i["users"] for i in combined_industries if i["sector"] == sid)
    should_be_here = [ind_id for ind_id, suggested in nav_mismatches.items() if suggested == sid]
    aa_tagged = sum(1 for aa in nav_aa if sid in aa.get("sectorIds",[]))
    fund_tagged = sum(1 for f in nav_fundings if sid in f.get("sector",[]))
    sector_summary.append({
        "id": sid,
        "name": sdata["name"],
        "orphaned": sid in orphaned_sectors,
        "industryCount": len(mapped_industries),
        "users": mapped_users,
        "xlsxUsers": sector_users.get(sdata["name"], 0),
        "aaTagged": aa_tagged,
        "fundTagged": fund_tagged,
        "missingIndustries": should_be_here,
        "missingIndustriesNames": [nav_industries[i]["name"] for i in should_be_here if i in nav_industries],
    })

# ============================================================
# Build anytime action reachability
# ============================================================
aa_reach = []
for aa in nav_aa:
    reached = []
    for ind in combined_industries:
        iids = aa.get("industryIds",[])
        sids = aa.get("sectorIds",[])
        if aa.get("applyToAllUsers"):
            reached.append({"id": ind["id"], "match": "universal"})
        elif ind["id"] in iids:
            reached.append({"id": ind["id"], "match": "industryId"})
        elif ind["sector"] in sids:
            reached.append({"id": ind["id"], "match": "sectorId"})
    aa_reach.append({
        "id": aa["id"],
        "name": aa["name"],
        "categories": aa.get("categories",[]),
        "industryIds": aa.get("industryIds",[]),
        "sectorIds": aa.get("sectorIds",[]),
        "applyToAllUsers": aa.get("applyToAllUsers", False),
        "reachCount": len(reached),
        "reached": reached,
    })

# ============================================================
# Phase analysis
# ============================================================
starting_phases = ["GUEST_MODE","GUEST_MODE_WITH_BUSINESS_STRUCTURE","NEEDS_BUSINESS_STRUCTURE","NEEDS_TO_FORM","FORMED"]
operate_starting = ["UP_AND_RUNNING"]
owning_phases = ["GUEST_MODE_OWNING","UP_AND_RUNNING_OWNING"]
operate_phases = operate_starting + owning_phases

sees_roadmap = sum(phases.get(p,0) for p in starting_phases + ["REMOTE_SELLER_WORKER","DOMESTIC_EMPLOYER"])
sees_operate = sum(phases.get(p,0) for p in operate_phases)
sees_aa_and_funding = sum(phases.get(p,0) for p in ["UP_AND_RUNNING","UP_AND_RUNNING_OWNING","GUEST_MODE_OWNING"])

phase_summary = {
    "totalBusinesses": total_businesses,
    "seesRoadmap": sees_roadmap,
    "seesOperate": sees_operate,
    "seesAAFunding": sees_aa_and_funding,
    "phases": [{"phase": p, "count": c, "group":
        "starting" if p in starting_phases else
        "operate" if p in operate_phases else
        "other"
    } for p, c in sorted(phases.items(), key=lambda x: -x[1])],
}

# ============================================================
# Plurmit inventory analysis
# ============================================================
PLURMIT_FILE = os.path.join(ROOT, "data", "plurmits.xlsx")
plurmit_data = None

if os.path.exists(PLURMIT_FILE):
    print(f"Reading plurmits: {PLURMIT_FILE}")
    pwb = openpyxl.load_workbook(PLURMIT_FILE, data_only=True)
    pws = pwb["Plurmits-All Plurmits"]
    pheaders = [c.value for c in next(pws.iter_rows(min_row=1, max_row=1))]
    prows = []
    for prow in pws.iter_rows(min_row=2, values_only=True):
        if prow[0]: prows.append(dict(zip(pheaders, prow)))

    # Agency DB integrations (from codebase analysis)
    api_integrations = [
        {"agency":"NJ Treasury (NJFAST)","agencyShort":"Treasury","type":"Submit+Pay","tasks":["form-business-entity"],"aas":[],"desc":"Business entity formation filing and payment"},
        {"agency":"NICUSA / NJ Treasury","agencyShort":"NICUSA","type":"Query","tasks":["search-business-name","search-business-name-nexus"],"aas":[],"desc":"Business name availability search"},
        {"agency":"DCA — License Status (Dynamics 365)","agencyShort":"DCA","type":"Query","tasks":["apply-for-shop-license","appraiser-company-register","authorization-architect-firm","authorization-landscape-architect-firm","cemetery-certificate","consulting-firm-headhunter-reg","electrologist-office-license","entertainment-agency-reg","funeral-registration","health-club-registration","home-health-aide-license","license-massage-therapy","moving-company-license","oos-pharmacy-registration","pharmacy-license","register-accounting-firm","register-home-contractor","search-licenses-employment-agency","telemarketing-license","ticket-broker-reseller-registration","temp-help-consulting-firm-combined-reg","temporary-help-service-firm-reg"],"aas":[],"desc":"22 DCA license status lookups via Dynamics 365"},
        {"agency":"DCA — Elevator Safety (Dynamics 365)","agencyShort":"DCA","type":"Query+Register","tasks":["elevator-registration"],"aas":[],"desc":"Elevator registration, inspection, violations"},
        {"agency":"DCA — Housing (Dynamics 365)","agencyShort":"DCA","type":"Query","tasks":["hotel-motel-registration","multiple-dwelling-registration"],"aas":[],"desc":"Hotel/motel and multiple dwelling registration status"},
        {"agency":"DEP — CRTK","agencyShort":"DEP","type":"Query","tasks":["community-right-to-know-survey"],"aas":[],"desc":"Community Right to Know facility lookup"},
        {"agency":"DEP — X-ray Registration","agencyShort":"DEP","type":"Query","tasks":["xray-reg"],"aas":[],"desc":"X-ray equipment registration status"},
        {"agency":"NJ Treasury — Cigarette License","agencyShort":"Treasury","type":"Submit+Pay","tasks":["cigarette-license"],"aas":[],"desc":"Cigarette license application and payment"},
        {"agency":"ABC (Alcoholic Beverage Control)","agencyShort":"ABC","type":"Submit","tasks":[],"aas":["emergency-trip-permit"],"desc":"Emergency trip permit application"},
        {"agency":"NJ Treasury — Tax Clearance","agencyShort":"Treasury","type":"Submit","tasks":[],"aas":["tax-clearance-certificate"],"desc":"Tax clearance certificate request"},
        {"agency":"NJ Treasury — Tax Filing","agencyShort":"Treasury","type":"Query","tasks":[],"aas":[],"desc":"Tax filing calendar and deadlines","dashboard":"Filings Calendar"},
        {"agency":"DOL (Dept. of Labor)","agencyShort":"DOL","type":"Query","tasks":[],"aas":[],"desc":"Employer contribution rates","dashboard":"Employer Rates"},
        {"agency":"MyNJ","agencyShort":"MyNJ","type":"Submit","tasks":[],"aas":[],"desc":"MyNJ account self-registration","dashboard":"Account Creation"},
        {"agency":"Power Automate (Email Relay)","agencyShort":"MSFT","type":"Send","tasks":["env-requirements","community-right-to-know-survey"],"aas":[],"desc":"Sends personalized email for env/CRTK results"},
    ]

    api_task_slugs = set()
    api_aa_slugs = set()
    for ai in api_integrations:
        api_task_slugs.update(ai["tasks"])
        api_aa_slugs.update(ai["aas"])

    def ps(val): return str(val or "").strip()

    biz_keywords = ["yes this is","yes businesses often","yes the shop owner","yes businesses use","yes, but primarily"]
    all_ind_ids = set(k for k,v in nav_industries.items() if v.get("isEnabled"))

    plurmits_out = []
    for pr in prows:
        pname = ps(pr.get("Plurmit"))
        dept = ps(pr.get("Department"))
        biz_app = any(kw in ps(pr.get("Does this plurmit apply to businesses?")).lower() for kw in biz_keywords)
        online = ps(pr.get("Online Status [Normalized]"))
        if online == "None": online = ""
        worth = ps(pr.get("Worth Integrating or Mentioning?"))
        integ = ps(pr.get("Is it Integrated or mentioned yet?"))
        cms = ps(pr.get("Consumer Affairs Licenses in CMS"))
        if cms == "None": cms = ""
        bnj_ind = ps(pr.get("Relates to B.NJ Industry"))
        if bnj_ind == "None": bnj_ind = ""
        ind_muni = ps(pr.get("Industry [MUNI]"))
        if ind_muni == "None": ind_muni = ""
        mlo = ps(pr.get("MLO Profession (using Boomi)"))
        if mlo == "None": mlo = ""
        notes = ps(pr.get("Notes on integration/mention analysis"))
        if notes == "None": notes = ""
        vol_raw = ps(pr.get("Annual volume of submissions"))
        vol = 0
        try: vol = int(float(str(vol_raw).replace(",","")))
        except: pass

        # Integration level
        if integ in ["Read and/or Write","Yes"]:
            int_level = "api"
        elif integ == "Read":
            int_level = "read"
        elif integ == "Mentioned":
            int_level = "mentioned"
        else:
            int_level = "none"

        # Match to navigator tasks
        matched_tasks = []
        if cms:
            for tn, sl in task_name_to_slug.items():
                if sl in cms.lower() or tn.lower() in cms.lower():
                    if sl not in matched_tasks: matched_tasks.append(sl)
            for sl in set(task_name_to_slug.values()):
                if sl in cms.lower() and sl not in matched_tasks:
                    matched_tasks.append(sl)

        # Match to industries
        matched_inds = []
        for fv in [bnj_ind, ind_muni]:
            if not fv: continue
            for iid in all_ind_ids:
                if iid in fv.lower() or iid.replace("-"," ") in fv.lower():
                    if iid not in matched_inds: matched_inds.append(iid)

        # Check API integration
        has_api = any(t in api_task_slugs for t in matched_tasks)

        # Priority
        if "High Priority" in worth: pri = "high"
        elif "Low Priority" in worth: pri = "low"
        elif "Mention" in worth: pri = "mention"
        elif worth == "No": pri = "no"
        else: pri = "unassessed"

        plurmits_out.append({
            "name": pname, "department": dept, "bizApplicable": biz_app,
            "onlineStatus": online, "integrationLevel": int_level, "priority": pri,
            "volume": vol, "matchedTasks": matched_tasks, "matchedIndustries": matched_inds,
            "hasApi": has_api, "cmsLink": cms[:120], "notes": notes[:200],
        })

    # Department summary
    dept_sum = {}
    for p in plurmits_out:
        d = p["department"]
        if d not in dept_sum:
            dept_sum[d] = {"name":d,"total":0,"bizApplicable":0,"mentioned":0,"read":0,"api":0,"none":0,"highPri":0,"lowPri":0,"online":0,"volume":0}
        ds = dept_sum[d]
        ds["total"] += 1; ds["volume"] += p["volume"]
        if p["bizApplicable"]: ds["bizApplicable"] += 1
        if p["onlineStatus"] in ["Available","Submitted"]: ds["online"] += 1
        ds[p["integrationLevel"]] += 1
        if p["priority"] == "high": ds["highPri"] += 1
        if p["priority"] == "low": ds["lowPri"] += 1

    # Industry -> plurmit count
    ind_plurmit_map = {}
    for p in plurmits_out:
        for iid in p["matchedIndustries"]:
            if iid not in ind_plurmit_map:
                ind_plurmit_map[iid] = {"total":0,"integrated":0,"api":0,"names":[]}
            ind_plurmit_map[iid]["total"] += 1
            if p["integrationLevel"] in ["mentioned","read","api"]: ind_plurmit_map[iid]["integrated"] += 1
            if p["hasApi"]: ind_plurmit_map[iid]["api"] += 1
            ind_plurmit_map[iid]["names"].append(p["name"])

    biz_only = [p for p in plurmits_out if p["bizApplicable"]]
    total_vol = sum(p["volume"] for p in plurmits_out)
    biz_vol = sum(p["volume"] for p in biz_only)
    api_pls = [p for p in biz_only if p["hasApi"] or p["integrationLevel"]=="api"]
    info_pls = [p for p in biz_only if p["matchedTasks"] and not p["hasApi"] and p["integrationLevel"]!="api"]
    read_pls = [p for p in biz_only if p["integrationLevel"]=="read" and not p["matchedTasks"] and not p["hasApi"]]
    ment_pls = [p for p in biz_only if p["integrationLevel"]=="mentioned" and not p["matchedTasks"]]
    none_pls = [p for p in biz_only if p["integrationLevel"]=="none" and not p["matchedTasks"] and not p["hasApi"]]

    # Compute actual BNJ engagement per category
    slug_eng = {}
    slug_comp = {}
    for tp_row in task_progress:
        tn = tp_row["task"]
        sl = task_name_to_slug.get(tn, tn)
        slug_eng[sl] = slug_eng.get(sl, 0) + tp_row["total"]
        slug_comp[sl] = slug_comp.get(sl, 0) + tp_row["completed"]

    def cat_engagement(permits, include_unmatched_api=False):
        seen = set()
        total_eng, total_comp = 0, 0
        for p in permits:
            for t in p["matchedTasks"]:
                if t not in seen:
                    seen.add(t)
                    total_eng += slug_eng.get(t, 0)
                    total_comp += slug_comp.get(t, 0)
        # For api-level permits without matched tasks, include engagement from
        # all API integration task slugs (e.g. formation, business name search)
        if include_unmatched_api:
            for ai in api_integrations:
                for t in ai["tasks"] + ai["aas"]:
                    if t not in seen:
                        seen.add(t)
                        total_eng += slug_eng.get(t, 0)
                        total_comp += slug_comp.get(t, 0)
        return total_eng, total_comp

    api_eng, api_comp = cat_engagement(api_pls, include_unmatched_api=True)
    info_eng, info_comp = cat_engagement(info_pls)
    read_eng, read_comp = cat_engagement(read_pls)
    ment_eng, ment_comp = cat_engagement(ment_pls)
    all_eng = api_eng + info_eng + read_eng + ment_eng
    all_comp = api_comp + info_comp + read_comp + ment_comp

    # Compute roadmap reach per category: how many of 64 industries include at
    # least one task from this category in their roadmap?
    def cat_reach(permits, include_unmatched_api=False):
        task_set = set()
        for p in permits:
            task_set.update(p["matchedTasks"])
        if include_unmatched_api:
            for ai in api_integrations:
                task_set.update(ai["tasks"])
                task_set.update(ai["aas"])
        # Count industries whose roadmap includes at least one of these tasks
        # (checking both differentiating and universal tasks)
        count = 0
        aa_list = nav.get("anytimeActions", [])
        for ind_id, ind_data in enabled_industries.items():
            all_tasks = set(ind_data.get("roadmapTaskNames", []))
            if all_tasks & task_set:
                count += 1
                continue
            # Also check AAs: does any AA in task_set reach this industry?
            for aa in aa_list:
                if aa["id"] in task_set:
                    if aa.get("applyToAllUsers"):
                        count += 1; break
                    elif ind_id in aa.get("industryIds", []):
                        count += 1; break
                    elif ind_data.get("defaultSectorId") in aa.get("sectorIds", []):
                        count += 1; break
        return count

    api_reach = cat_reach(api_pls, include_unmatched_api=True)
    info_reach = cat_reach(info_pls)

    # Compute user reach per category (total users across industries that see these tasks)
    def cat_user_reach(permits, include_unmatched_api=False):
        task_set = set()
        for p in permits:
            task_set.update(p["matchedTasks"])
        if include_unmatched_api:
            for ai in api_integrations:
                task_set.update(ai["tasks"])
                task_set.update(ai["aas"])
        aa_list = nav.get("anytimeActions", [])
        matched_inds = set()
        for ind_id, ind_data in enabled_industries.items():
            all_tasks = set(ind_data.get("roadmapTaskNames", []))
            if all_tasks & task_set:
                matched_inds.add(ind_id)
                continue
            for aa in aa_list:
                if aa["id"] in task_set:
                    if aa.get("applyToAllUsers"):
                        matched_inds.add(ind_id); break
                    elif ind_id in aa.get("industryIds", []):
                        matched_inds.add(ind_id); break
                    elif ind_data.get("defaultSectorId") in aa.get("sectorIds", []):
                        matched_inds.add(ind_id); break
        users = sum(industry_users.get(enabled_industries[iid]["name"].strip(), 0) for iid in matched_inds)
        return users

    api_users = cat_user_reach(api_pls, include_unmatched_api=True)
    info_users = cat_user_reach(info_pls)

    hi_gaps = [{"name":p["name"],"department":p["department"],"volume":p["volume"],"notes":p["notes"]}
               for p in sorted(biz_only, key=lambda x:-x["volume"]) if p["priority"]=="high" and p["integrationLevel"]=="none"]

    plurmit_data = {
        "coverage": {
            "total": len(plurmits_out),
            "totalVolume": total_vol,
            "bizApplicable": len(biz_only),
            "bizVolume": biz_vol,
            "bizApi": len(api_pls),
            "bizApiVol": sum(p["volume"] for p in api_pls),
            "bizApiEng": api_eng, "bizApiComp": api_comp, "bizApiReach": api_reach, "bizApiUsers": api_users,
            "bizInfo": len(info_pls),
            "bizInfoVol": sum(p["volume"] for p in info_pls),
            "bizInfoEng": info_eng, "bizInfoComp": info_comp, "bizInfoReach": info_reach, "bizInfoUsers": info_users,
            "bizRead": len(read_pls),
            "bizReadVol": sum(p["volume"] for p in read_pls),
            "bizReadEng": read_eng, "bizReadComp": read_comp,
            "bizMentioned": len(ment_pls),
            "bizMentionedVol": sum(p["volume"] for p in ment_pls),
            "bizMentionedEng": ment_eng, "bizMentionedComp": ment_comp,
            "bizNone": len(none_pls),
            "bizNoneVol": sum(p["volume"] for p in none_pls),
            "bizAllEng": all_eng, "bizAllComp": all_comp,
            "bizHighPri": sum(1 for p in biz_only if p["priority"]=="high"),
            "bizLowPri": sum(1 for p in biz_only if p["priority"]=="low"),
            "highPriGaps": hi_gaps,
        },
        "departments": sorted(dept_sum.values(), key=lambda x:-x["total"]),
        "apiIntegrations": api_integrations,
        "apiTaskSlugs": sorted(api_task_slugs),
        "apiTaskNames": sorted(set(tn for tn, sl in task_name_to_slug.items() if sl in api_task_slugs)),
        "apiAASlugs": sorted(api_aa_slugs),
        "industryPlurmitMap": ind_plurmit_map,
        "plurmits": [{"name":p["name"],"dept":p["department"],"biz":p["bizApplicable"],
                      "online":p["onlineStatus"],"level":p["integrationLevel"],"pri":p["priority"],
                      "vol":p["volume"],"tasks":p["matchedTasks"],"inds":p["matchedIndustries"],
                      "api":p["hasApi"]} for p in biz_only],
    }
    print(f"  Plurmits: {len(plurmits_out)} total, {len(biz_only)} business-applicable")
    print(f"  API-connected tasks: {len(api_task_slugs)}, API-connected AAs: {len(api_aa_slugs)}")

# ============================================================
# Task trigger mapping (how to reach each add-on task)
# ============================================================
import glob as _glob

ADDON_DIR = os.path.join(ROOT, "data", "add-ons")
NEQ_FILE = os.path.join(ROOT, "data", "nonEssentialQuestions.json")

# Try loading from navigator repo if available, fall back to cached copies
nav_content_addons = os.path.join(ROOT, "..", "nav", "content", "src", "roadmaps", "add-ons")
nav_content_neq = os.path.join(ROOT, "..", "nav", "content", "src", "roadmaps", "nonEssentialQuestions.json")

addon_source = nav_content_addons if os.path.isdir(nav_content_addons) else ADDON_DIR
neq_source = nav_content_neq if os.path.isfile(nav_content_neq) else NEQ_FILE

task_triggers = {}
codebase_neqs = []
industry_neq_tasks = {}

if os.path.isdir(addon_source):
    # Load add-on -> task mapping
    addon_to_tasks = {}
    for f in _glob.glob(os.path.join(addon_source, "*.json")):
        aname = os.path.basename(f).replace(".json","")
        adata = json.load(open(f))
        atasks = []
        for step in adata.get("roadmapSteps", []):
            t = step.get("task") or step.get("licenseTask")
            if t: atasks.append(t)
        addon_to_tasks[aname] = atasks

    # Load NEQ -> addon mapping
    neq_addon_map = {}
    if os.path.isfile(neq_source):
        neq_list = json.load(open(neq_source)).get("nonEssentialQuestionsArray", [])
        for nq in neq_list:
            if nq.get("addOnWhenYes"):
                raw_q = nq["questionText"].replace("`","")
                # Strip markdown links [text](url) -> text
                import re as _re
                clean_q = _re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', raw_q).strip()
                neq_addon_map[nq["addOnWhenYes"]] = {"neqId":nq["id"],"q":clean_q,"answer":"Yes"}
            if nq.get("addOnWhenNo"):
                raw_q = nq["questionText"].replace("`","")
                import re as _re
                clean_q = _re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', raw_q).strip()
                neq_addon_map[nq["addOnWhenNo"]] = {"neqId":nq["id"],"q":clean_q,"answer":"No"}

    # Build NEQ id -> list of industry IDs that show this question
    neq_ind_map = {}  # neqId -> [industryId, ...]
    ind_source = addon_source.replace("add-ons", "industries")
    if os.path.isdir(ind_source):
        for f in _glob.glob(os.path.join(ind_source, "*.json")):
            idata = json.load(open(f))
            if not idata.get("isEnabled"): continue
            for neq_id in idata.get("nonEssentialQuestionsIds", []):
                if neq_id not in neq_ind_map:
                    neq_ind_map[neq_id] = []
                neq_ind_map[neq_id].append(idata["id"])

    # Profile/legal structure triggers (from buildUserRoadmap.ts analysis)
    hardcoded_triggers = {
        "permanent-location-business":{"t":"profile","d":"In your Profile under 'Location-Based Requirements', answer 'No' to 'Is this a home-based business?' (only shown for 47 industries where canBeHomeBased is true)"},
        "liquor-license":{"t":"profile","d":"In your Profile, answer 'Yes' to 'Will you need a liquor license?' (only shown for industries with liquor license question)"},
        "cpa":{"t":"profile","d":"In your Profile, answer 'Yes' to 'Does your business require a CPA?' (accounting industry)"},
        "petcare-license":{"t":"profile","d":"In your Profile, answer 'Yes' to 'Will you house animals?' (pet care industry)"},
        "home-based-transportation":{"t":"profile","d":"In your Profile, answer 'Yes' to 'Is this a home-based business?' (transportation industries only)"},
        "planned-renovation":{"t":"profile","d":"In your Profile, answer 'No' to home-based, then 'Yes' to 'Are you planning renovations?' (only for non-home-based businesses)"},
        "cannabis-annual":{"t":"profile","d":"Select 'Annual' cannabis license type"},
        "cannabis-conditional":{"t":"profile","d":"Select 'Conditional' cannabis license type"},
        "real-estate-appraisal-management":{"t":"profile","d":"Answer 'Yes' — is this an appraisal management company?"},
        "real-estate-appraiser":{"t":"profile","d":"Answer 'No' — not an appraisal management company"},
        "car-service-standard":{"t":"profile","d":"Select 'Standard' or 'Both' for car service type"},
        "car-service-high-capacity":{"t":"profile","d":"Select 'High capacity' or 'Both' for car service type"},
        "public-works-contractor":{"t":"profile","d":"Answer 'Yes' to 'Are you a public works contractor?'"},
        "employment-agency-job-seekers":{"t":"profile","d":"Select 'Job seekers' for employment personnel service type"},
        "employment-agency-employers-temporary":{"t":"profile","d":"Select 'Employers' → 'Temporary' placement type"},
        "employment-agency-employers-permanent":{"t":"profile","d":"Select 'Employers' → 'Permanent' placement type"},
        "employment-agency-employers-both":{"t":"profile","d":"Select 'Employers' → 'Both' placement type"},
        "interstate-logistics":{"t":"profile","d":"Answer 'Yes' to 'Will you transport goods interstate?'"},
        "interstate-moving":{"t":"profile","d":"Answer 'Yes' to 'Will you move goods interstate?'"},
        "elevator-owning-business":{"t":"profile","d":"Answer 'Yes' to 'Does your business own elevators?'"},
        "daycare":{"t":"profile","d":"Answer 'Yes' to 'Will you care for 6 or more children?'"},
        "family-daycare":{"t":"profile","d":"Answer 'No' to 'Will you care for 6 or more children?'"},
        "will-sell-pet-care-items":{"t":"profile","d":"Answer 'Yes' to 'Will you sell pet care items?'"},
        "residential-landlord-long-term-many-units":{"t":"profile","d":"Select 'Long-term' or 'Both' lease type, then 'Yes' to 3+ rental units"},
        "residential-landlord-long-term-few-units":{"t":"profile","d":"Select 'Long-term' or 'Both' lease type, then 'No' to 3+ rental units"},
        "short-term-rental-registration":{"t":"profile","d":"Select 'Short-term rental' or 'Both' for property lease type"},
        "construction-home-renovation":{"t":"profile","d":"Select 'Home renovations' or 'Both' for construction type"},
        "construction-new-home-construction":{"t":"profile","d":"Select 'New home construction' or 'Both' for construction type"},
        "env-requirements":{"t":"profile","d":"Select 'All Other Businesses' industry AND answer 'No' to home-based"},
        "logistics-modification":{"t":"auto","d":"Automatically applied for Logistics industry"},
        "permanent-location-business-landlord":{"t":"auto","d":"Automatically applied for Residential Landlord industry"},
        "public-record-filing":{"t":"legalStructure","d":"Select LLC, C-Corp, S-Corp, LP, LLP, or Nonprofit as your business structure"},
        "public-record-filing-foreign":{"t":"legalStructure","d":"Foreign business with a legal structure requiring public filing"},
        "trade-name":{"t":"legalStructure","d":"Select Sole Proprietorship or General Partnership as your business structure"},
        "llc":{"t":"legalStructure","d":"Select LLC as your business structure"},
        "scorp":{"t":"legalStructure","d":"Select S-Corporation as your business structure"},
        "nonprofit":{"t":"legalStructure","d":"Select Nonprofit as your business structure"},
        "nonprofit-and-corp-foreign":{"t":"legalStructure","d":"Foreign business with S-Corp, C-Corp, or Nonprofit structure"},
        "raffle-bingo-games":{"t":"legalStructure+profile","d":"Select Nonprofit, then answer 'Yes' to 'Will your nonprofit hold raffle or bingo games?' in your Profile"},
        "business-vehicle":{"t":"roadmapTask","d":"Click 'Manage Business Vehicles' in your roadmap and indicate you manage business vehicles"},
        "foreign-remote-worker":{"t":"persona","d":"Select 'Foreign business' → indicate you have employees in NJ"},
        "foreign-remote-seller":{"t":"persona","d":"Select 'Foreign business' → indicate you have revenue/transactions in NJ"},
        "foreign-nexus":{"t":"persona","d":"Select 'Foreign business' → indicate you have office/property/employees/vehicles in NJ"},
        "oos-pharmacy":{"t":"persona+industry","d":"Foreign nexus business in Pharmacy industry"},
    }

    # Dynamically resolve which industries show each profile question
    # by reading industryOnboardingQuestions from industry JSON files
    ioq_industries = {}  # ioq_field -> [industry_ids]
    can_have_perm = []   # industries with canHavePermanentLocation
    can_be_home = []     # industries with canBeHomeBased
    if os.path.isdir(ind_source):
        for f in _glob.glob(os.path.join(ind_source, "*.json")):
            idata = json.load(open(f))
            if not idata.get("isEnabled"): continue
            if idata.get("canHavePermanentLocation"):
                can_have_perm.append(idata["id"])
            ioq = idata.get("industryOnboardingQuestions", {})
            for field, val in ioq.items():
                if val:
                    if field not in ioq_industries:
                        ioq_industries[field] = []
                    ioq_industries[field].append(idata["id"])
            if ioq.get("canBeHomeBased"):
                can_be_home.append(idata["id"])

    # Map add-on names to their applicable industry lists
    addon_industry_map = {
        "permanent-location-business": [i for i in can_have_perm if i in can_be_home],
        "planned-renovation": [i for i in can_have_perm if i in can_be_home],
        "home-based-transportation": ioq_industries.get("isTransportation", []),
        "elevator-owning-business": can_have_perm,
        "liquor-license": ioq_industries.get("isLiquorLicenseApplicable", []),
        "cpa": ioq_industries.get("isCpaRequiredApplicable", []),
        "petcare-license": ioq_industries.get("isPetCareHousingApplicable", []),
        "will-sell-pet-care-items": ioq_industries.get("willSellPetCareItems", []),
        "daycare": ioq_industries.get("isChildcareForSixOrMore", []),
        "family-daycare": ioq_industries.get("isChildcareForSixOrMore", []),
        "cannabis-annual": ioq_industries.get("isCannabisLicenseTypeApplicable", []),
        "cannabis-conditional": ioq_industries.get("isCannabisLicenseTypeApplicable", []),
        "real-estate-appraisal-management": ioq_industries.get("isRealEstateAppraisalManagementApplicable", []),
        "real-estate-appraiser": ioq_industries.get("isRealEstateAppraisalManagementApplicable", []),
        "car-service-standard": ioq_industries.get("isCarServiceApplicable", []),
        "car-service-high-capacity": ioq_industries.get("isCarServiceApplicable", []),
        "employment-agency-job-seekers": ioq_industries.get("isEmploymentAndPersonnelTypeApplicable", []),
        "employment-agency-employers-temporary": ioq_industries.get("isEmploymentAndPersonnelTypeApplicable", []),
        "employment-agency-employers-permanent": ioq_industries.get("isEmploymentAndPersonnelTypeApplicable", []),
        "employment-agency-employers-both": ioq_industries.get("isEmploymentAndPersonnelTypeApplicable", []),
        "interstate-logistics": ioq_industries.get("isInterstateLogisticsApplicable", []),
        "interstate-moving": ioq_industries.get("isInterstateMovingApplicable", []),
        "construction-home-renovation": ioq_industries.get("isConstructionTypeApplicable", []),
        "construction-new-home-construction": ioq_industries.get("isConstructionTypeApplicable", []),
        "public-works-contractor": can_have_perm,  # shown to all permanent location industries
        "residential-landlord-long-term-many-units": ioq_industries.get("canHaveThreeOrMoreRentalUnits", []),
        "residential-landlord-long-term-few-units": ioq_industries.get("canHaveThreeOrMoreRentalUnits", []),
        "short-term-rental-registration": ioq_industries.get("whatIsPropertyLeaseType", []),
        "env-requirements": ["generic"],
        "logistics-modification": ["logistics"],
        "permanent-location-business-landlord": ["residential-landlord"],
    }

    # Step names for display
    step_names = {1:"Plan Your Business",2:"Register Your Business",3:"After Registering Your Business",4:"Before Opening Your Site"}

    # Build task_slug -> triggers
    for addon_name, atasks in addon_to_tasks.items():
        # Get the step numbers for each task from the add-on JSON
        addon_path = os.path.join(addon_source, f"{addon_name}.json")
        addon_json = json.load(open(addon_path)) if os.path.isfile(addon_path) else {}
        task_step_map = {}
        for rs in addon_json.get("roadmapSteps", []):
            t = rs.get("task") or rs.get("licenseTask")
            if t: task_step_map[t] = rs.get("step")

        for task_slug in atasks:
            if task_slug not in task_triggers:
                task_triggers[task_slug] = []
            trigger = {"addon": addon_name}
            if addon_name in neq_addon_map:
                ni = neq_addon_map[addon_name]
                trigger["type"] = "neq"
                trigger["neqId"] = ni["neqId"]
                trigger["question"] = ni["q"]
                trigger["answer"] = ni["answer"]
                trigger["detail"] = f"Answer '{ni['answer']}' to: \"{ni['q'][:120]}\""
                trigger["industries"] = neq_ind_map.get(ni["neqId"], [])
            elif addon_name in hardcoded_triggers:
                ht = hardcoded_triggers[addon_name]
                trigger["type"] = ht["t"]
                trigger["detail"] = ht["d"]
                # Add applicable industries from the dynamic map
                if addon_name in addon_industry_map:
                    trigger["industries"] = addon_industry_map[addon_name]
            else:
                trigger["type"] = "unknown"
                trigger["detail"] = f"Triggered by add-on '{addon_name}'"
            task_triggers[task_slug].append(trigger)
            # Add step info
            step_num = task_step_map.get(task_slug)
            if step_num and task_slug in task_triggers:
                for tr in task_triggers[task_slug]:
                    if tr.get("addon") == addon_name and "step" not in tr:
                        tr["step"] = step_num
                        tr["stepName"] = step_names.get(step_num, f"Step {step_num}")

    print(f"  Task triggers: {len(task_triggers)} add-on tasks mapped")

    # ================================================================
    # Build complete NEQ data from codebase (all 56, not just XLSX 22)
    # ================================================================
    codebase_neqs = []
    if os.path.isfile(neq_source):
        neq_list = json.load(open(neq_source)).get("nonEssentialQuestionsArray", [])
        for nq in neq_list:
            raw_q = nq.get("questionText", "")
            # Clean markdown from question text
            clean_q = raw_q.replace("`", "").split("|")[0] if "`" in raw_q else raw_q
            clean_q = clean_q.replace("[", "").replace("](", " ").split(")")[0] if "[" in clean_q else clean_q
            clean_q = clean_q.strip().strip('"').strip()
            
            neq_id = nq["id"]
            yes_addon = nq.get("addOnWhenYes", "")
            no_addon = nq.get("addOnWhenNo", "")
            
            # Find downstream tasks for each answer path
            yes_tasks = []
            no_tasks = []
            for addon_name, task_list in [(yes_addon, yes_tasks), (no_addon, no_tasks)]:
                if not addon_name: continue
                for tslug, trigs in task_triggers.items():
                    if any(t.get("addon") == addon_name for t in trigs):
                        task_list.append(tslug)
            
            # Which industries show this NEQ?
            industries = neq_ind_map.get(neq_id, [])
            
            # Calculate downstream engagement
            all_tasks = sorted(set(yes_tasks + no_tasks))
            total_eng = sum(slug_eng_map.get(ts, 0) for ts in all_tasks) if 'slug_eng_map' in dir() else 0
            
            codebase_neqs.append({
                "id": neq_id,
                "question": clean_q[:200],
                "industries": industries,
                "industryCount": len(industries),
                "yesAddon": yes_addon,
                "noAddon": no_addon,
                "yesTasks": yes_tasks,
                "noTasks": no_tasks,
                "allTasks": all_tasks,
                "taskCount": len(all_tasks),
            })
        print(f"  Codebase NEQs: {len(codebase_neqs)} (triggering tasks for {sum(1 for n in codebase_neqs if n['taskCount']>0)} of them)")

    # Build per-industry NEQ-triggered tasks
    industry_neq_tasks = {}  # ind_id -> set of task slugs reachable via NEQs
    if os.path.isdir(ind_source):
        for f in _glob.glob(os.path.join(ind_source, "*.json")):
            idata = json.load(open(f))
            if not idata.get("isEnabled"): continue
            ind_id = idata["id"]
            neq_ids = idata.get("nonEssentialQuestionsIds", [])
            neq_task_set = set()
            for nq_id in neq_ids:
                # Find this NEQ in our codebase list
                nq_entry = next((n for n in codebase_neqs if n["id"] == nq_id), None)
                if nq_entry:
                    neq_task_set.update(nq_entry["allTasks"])
            industry_neq_tasks[ind_id] = neq_task_set

    # Patch combined_industries with NEQ task data
    for ci in combined_industries:
        ind_id = ci["id"]
        neq_tasks = industry_neq_tasks.get(ind_id, set())
        ci["neqTaskCount"] = len(neq_tasks)
        ci["neqTasks"] = sorted(neq_tasks)
        # "Truly unique content" = base diff tasks + NEQ-triggered tasks
        ci["totalContentTasks"] = ci["totalDiffTasks"] + len(neq_tasks)
        # NEQ-only industries: zero base diff tasks but have NEQ tasks
        ci["uniqueViaNeqOnly"] = ci["totalDiffTasks"] == 0 and len(neq_tasks) > 0

    # Also add step numbers for base industry tasks (not add-ons)
    ind_dir = os.path.join(addon_source.replace("add-ons",""), "industries")
    if os.path.isdir(ind_dir):
        base_task_steps = {}  # task_slug -> set of step numbers
        for f in _glob.glob(os.path.join(ind_dir, "*.json")):
            idata = json.load(open(f))
            if not idata.get("isEnabled"): continue
            for rs in idata.get("roadmapSteps", []):
                t = rs.get("task") or rs.get("licenseTask")
                if t:
                    if t not in base_task_steps:
                        base_task_steps[t] = set()
                    base_task_steps[t].add(rs.get("step"))
        # Store in task_triggers for tasks NOT already there
        for tslug, step_nums in base_task_steps.items():
            if tslug not in task_triggers:
                task_triggers[tslug] = []
            # Add a "base" entry if no triggers exist
            if not any(tr.get("type") for tr in task_triggers.get(tslug, [])):
                sn = min(step_nums) if step_nums else None
                task_triggers[tslug] = [{"type":"base","addon":None,
                    "detail":"Base industry roadmap task — included automatically for this industry",
                    "step":sn,"stepName":step_names.get(sn, f"Step {sn}") if sn else None}]
            # Ensure all entries have step info
            for tr in task_triggers.get(tslug, []):
                if "step" not in tr and step_nums:
                    tr["step"] = min(step_nums)
                    tr["stepName"] = step_names.get(min(step_nums), f"Step {min(step_nums)}")

# ============================================================
# Enrich NEQ data with downstream tasks and engagement
# ============================================================
# Map analytics question names to profile fields and their add-ons
question_addon_map = {
    "Home-Based Business": {"field":"homeBasedBusiness","yes":["home-based-transportation"],"no":["permanent-location-business","planned-renovation"],
        "note":"No → adds 6 location tasks (evaluate, site requirements, lease, zoning, safety permits, mercantile). Yes + transportation → adds trucking parking."},
    "Provides Staffing Service": {"field":"providesStaffingService","yes":[],"no":[],
        "note":"If true, overrides industryId to 'employment-agency' — changes entire roadmap, not just add-ons."},
    "Interstate Logistics": {"field":"interstateLogistics","yes":["interstate-logistics"],"no":[],
        "note":"Adds BOC-3, USDOT, IRP, and IFTA tasks for interstate freight."},
    "Childcare for 6+ Children": {"field":"isChildcareForSixOrMore","yes":["daycare"],"no":["family-daycare"],
        "note":"Yes → daycare center path (evaluate location, site requirements, license). No → family daycare registration."},
    "Liquor License": {"field":"liquorLicense","yes":["liquor-license"],"no":[],
        "note":"Adds liquor license availability confirmation task."},
    "Requires CPA": {"field":"requiresCpa","yes":["cpa"],"no":[],
        "note":"Adds public accountant license task."},
    "Pet Care Housing": {"field":"petCareHousing","yes":["petcare-license"],"no":[],
        "note":"Adds pet care facility license task."},
    "Public Works Contractor": {"field":"publicWorksContractor","yes":["public-works-contractor"],"no":[],
        "note":"Adds public works contractor registration (PWCR) task."},
    "Sells Pet Care Items": {"field":"willSellPetCareItems","yes":["will-sell-pet-care-items"],"no":[],
        "note":"Adds retail license search task."},
    "Interstate Moving": {"field":"interstateMoving","yes":["interstate-moving"],"no":[],
        "note":"Adds BOC-3, USDOT, IRP, and IFTA tasks for interstate moving."},
    "Owns Elevators": {"field":"elevatorOwningBusiness","yes":["elevator-owning-business"],"no":[],
        "note":"Adds elevator registration task (DCA Dynamics API integration)."},
    "Cannabis Microbusiness": {"field":"cannabisMicrobusiness","yes":[],"no":[],
        "note":"Stored in profile but does not trigger add-ons. Display purposes only."},
    "Has 3+ Rental Units": {"field":"hasThreeOrMoreRentalUnits","yes":["residential-landlord-long-term-many-units"],"no":["residential-landlord-long-term-few-units"],
        "note":"Yes → multiple dwelling registration. No → landlord registration + smoke detector certificate."},
    "Interstate Transport": {"field":"interstateLogistics","yes":["interstate-logistics"],"no":[],
        "note":"Same trigger as Interstate Logistics — shown to different industries."},
    "Certified Interior Designer": {"field":"certifiedInteriorDesigner","yes":[],"no":[],
        "note":"If false, overrides industryId from 'interior-designer' to 'generic'. Changes entire roadmap."},
    "Real Estate Appraisal Management": {"field":"realEstateAppraisalManagement","yes":["real-estate-appraisal-management"],"no":["real-estate-appraiser"],
        "note":"Yes → appraiser certification + AMC registration. No → individual appraiser license."},
    "Raffle/Bingo Games": {"field":"raffleBingoGames","yes":["raffle-bingo-games"],"no":[],
        "note":"Only asked if legalStructure = nonprofit. Adds raffle/bingo game license task."},
    "Car Service": {"field":"carService","yes":["car-service-standard","car-service-high-capacity"],"no":[],
        "note":"Standard → taxi insurance, local auth, driver certification. High capacity → transport insurance, inspection, CPCN, USDOT."},
    "Owns Carnival Rides": {"field":"carnivalRideOwningBusiness","yes":[],"no":[],
        "note":"Triggers anytime actions only (carnival ride modification, operating fire permit)."},
    "Traveling Circus/Carnival": {"field":"travelingCircusOrCarnivalOwningBusiness","yes":[],"no":[],
        "note":"Triggers anytime action only (operating carnival fire permit)."},
    "Vacant Property Owner": {"field":"vacantPropertyOwner","yes":[],"no":[],
        "note":"Triggers anytime action only (vacant building fire permit)."},
    "Open 2+ Years": {"field":"openTwoOrMoreYears","yes":[],"no":[],
        "note":"No add-on or anytime action mapped. Appears unused."},
}

# Build slug -> engagement lookup
slug_eng_map = {}
for tp_row in task_progress:
    sl = tp_row.get("slug") or task_name_to_slug.get(tp_row["task"], tp_row["task"])
    slug_eng_map[sl] = slug_eng_map.get(sl, 0) + tp_row["total"]

for nq in neq_data:
    qname = nq["question"]
    qmap = question_addon_map.get(qname)
    if not qmap:
        nq["downstream"] = None
        continue
    
    all_addons = qmap["yes"] + qmap["no"]
    downstream_tasks = set()
    for addon in all_addons:
        for task_slug, trigs in task_triggers.items():
            if any(t.get("addon") == addon for t in trigs):
                downstream_tasks.add(task_slug)
    
    total_eng = sum(slug_eng_map.get(ts, 0) for ts in downstream_tasks)
    
    nq["downstream"] = {
        "field": qmap["field"],
        "note": qmap["note"],
        "tasks": sorted(downstream_tasks),
        "taskCount": len(downstream_tasks),
        "engagement": total_eng,
    }

# Track XLSX industries that didn't match any navigator industry
# First, enrich codebase NEQs with engagement data (slug_eng_map now exists)
for cnq in codebase_neqs:
    cnq["engagement"] = sum(slug_eng_map.get(ts, 0) for ts in cnq["allTasks"])

matched_names = set(ind["name"].strip() for ind in nav_industries.values() if ind.get("isEnabled"))
unmatched_xlsx = []
for name, cnt in industry_users.items():
    if name not in matched_names:
        unmatched_xlsx.append({"name": name, "users": cnt})

# ============================================================
# GA4 Analytics Processing
# ============================================================
import csv as _csv
import re as _re2

GA4_DIR = os.path.join(ROOT, "data")
ga4_data = None

def parse_ga4_csv(filepath):
    """Parse GA4 CSV, skipping comment lines starting with #"""
    if not os.path.isfile(filepath):
        return []
    rows = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        lines = [l for l in f if not l.startswith("#")]
    reader = _csv.DictReader(lines)
    for r in reader:
        rows.append(r)
    return rows

def safe_int(val):
    """Parse integer from GA4 CSV value, handling commas and empty strings"""
    if not val: return 0
    return int(str(val).replace(",","").strip())

def safe_float(val):
    """Parse float from GA4 CSV value"""
    if not val: return 0.0
    return float(str(val).replace(",","").strip())

ga4_pages_file = os.path.join(GA4_DIR, "ga4_page_views.csv")
ga4_events_file = os.path.join(GA4_DIR, "ga4_events.csv")
ga4_traffic_file = os.path.join(GA4_DIR, "ga4_traffic_sources.csv")
ga4_landing_file = os.path.join(GA4_DIR, "ga4_landing_pages.csv")

if os.path.isfile(ga4_pages_file):
    print("\nProcessing GA4 data...")

    # --- Parse date range from header ---
    ga4_date_range = {}
    with open(ga4_pages_file, "r") as f:
        for line in f:
            if "Start date" in line:
                d = line.strip().split(":")[-1].strip()
                ga4_date_range["start"] = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
            elif "End date" in line:
                d = line.strip().split(":")[-1].strip()
                ga4_date_range["end"] = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
            elif not line.startswith("#"):
                break

    # --- Page Views ---
    page_rows = parse_ga4_csv(ga4_pages_file)

    # --- Build urlSlug → taskId mapping from task markdown files ---
    url_to_id = {}
    task_content_dirs = [
        os.path.join(addon_source.replace("add-ons",""), "tasks"),
        os.path.join(addon_source.replace("add-ons",""), "license-tasks"),
        os.path.join(addon_source.replace("add-ons",""), "env-tasks"),
    ]
    for tcd in task_content_dirs:
        if not os.path.isdir(tcd): continue
        for f in _glob.glob(os.path.join(tcd, "*.md")):
            content = open(f).read()
            parts = content.split("---")
            if len(parts) < 3: continue
            fm = parts[1]
            task_id = None
            url_slug = None
            for line in fm.split("\n"):
                line = line.strip()
                if line.startswith("id:"):
                    val = line[3:].strip().strip('"').strip("'")
                    if val: task_id = val
                elif line.startswith("urlSlug:"):
                    val = line[8:].strip().strip('"').strip("'")
                    if val: url_slug = val
            if task_id and url_slug:
                url_to_id[url_slug] = task_id
    if url_to_id:
        print(f"  URL slug → task ID mappings: {len(url_to_id)} ({sum(1 for u,i in url_to_id.items() if u!=i)} differ)")

    # --- Page Views: match using urlSlug → taskId ---
    task_page_views = {}  # task_id -> {views, users, avgEngTime}
    for r in page_rows:
        path = r.get("Page path and screen class", "")
        if path.startswith("/tasks/"):
            url_slug = path.replace("/tasks/", "").strip("/").split("?")[0]
            if not url_slug or url_slug == "undefined": continue
            # Resolve URL slug to task ID
            task_id = url_to_id.get(url_slug, url_slug)
            views = safe_int(r.get("Views"))
            users = safe_int(r.get("Active users"))
            avg_time = safe_float(r.get("Average engagement time per active user"))
            if task_id not in task_page_views:
                task_page_views[task_id] = {"views": 0, "users": 0, "avgEngTime": 0}
            task_page_views[task_id]["views"] += views
            task_page_views[task_id]["users"] += users
            if avg_time > task_page_views[task_id]["avgEngTime"]:
                task_page_views[task_id]["avgEngTime"] = round(avg_time, 1)

    # Match anytime action pages: /actions/{slug}
    # Build AA urlSlug → id mapping
    aa_url_to_id = {}
    aa_content_dir = os.path.join(addon_source.replace("add-ons","").replace("roadmaps/",""), "anytime-action-tasks")
    if os.path.isdir(aa_content_dir):
        for f in _glob.glob(os.path.join(aa_content_dir, "*.md")):
            content = open(f).read()
            parts = content.split("---")
            if len(parts) < 3: continue
            fm = parts[1]
            aa_id = None
            aa_url = None
            for line in fm.split("\n"):
                line = line.strip()
                if line.startswith("id:"):
                    aa_id = line[3:].strip().strip('"').strip("'")
                elif line.startswith("urlSlug:"):
                    aa_url = line[8:].strip().strip('"').strip("'")
            if aa_id and aa_url:
                aa_url_to_id[aa_url] = aa_id

    aa_page_views = {}
    for r in page_rows:
        path = r.get("Page path and screen class", "")
        if path.startswith("/actions/"):
            url_slug = path.replace("/actions/", "").strip("/").split("?")[0]
            if not url_slug: continue
            aa_id = aa_url_to_id.get(url_slug, url_slug)
            views = safe_int(r.get("Views"))
            users = safe_int(r.get("Active users"))
            if aa_id not in aa_page_views:
                aa_page_views[aa_id] = {"views": 0, "users": 0}
            aa_page_views[aa_id]["views"] += views
            aa_page_views[aa_id]["users"] += users

    # Top pages overall
    top_pages = []
    for r in sorted(page_rows, key=lambda x: -safe_int(x.get("Views"))):
        path = r.get("Page path and screen class", "")
        if not path: continue
        top_pages.append({
            "path": path,
            "views": safe_int(r.get("Views")),
            "users": safe_int(r.get("Active users")),
            "avgEngTime": round(safe_float(r.get("Average engagement time per active user")), 1),
        })
        if len(top_pages) >= 50:
            break

    # Page category summary
    page_categories = {}
    for r in page_rows:
        path = r.get("Page path and screen class", "")
        parts = path.strip("/").split("/")
        prefix = "/" + parts[0] if parts[0] else "/"
        if prefix not in page_categories:
            page_categories[prefix] = {"pages": 0, "views": 0, "users": 0}
        page_categories[prefix]["pages"] += 1
        page_categories[prefix]["views"] += safe_int(r.get("Views"))
        page_categories[prefix]["users"] += safe_int(r.get("Active users"))
    page_cat_list = sorted(page_categories.items(), key=lambda x: -x[1]["views"])
    page_cat_list = [{"prefix": k, **v} for k, v in page_cat_list[:30]]

    # --- Events ---
    event_rows = parse_ga4_csv(ga4_events_file)
    events_summary = []
    for r in event_rows:
        name = r.get("Event name", "")
        if not name: continue
        events_summary.append({
            "event": name,
            "count": safe_int(r.get("Event count")),
            "users": safe_int(r.get("Total users")),
        })

    # Key funnel metrics from events
    event_lookup = {e["event"]: e for e in events_summary}
    funnel = {
        "pageViews": event_lookup.get("page_view", {}).get("count", 0),
        "firstVisit": event_lookup.get("first_visit", {}).get("count", 0),
        "firstVisitUsers": event_lookup.get("first_visit", {}).get("users", 0),
        "sessionStart": event_lookup.get("session_start", {}).get("count", 0),
        "onboardingStep": event_lookup.get("onboarding_step", {}).get("count", 0),
        "onboardingStepUsers": event_lookup.get("onboarding_step", {}).get("users", 0),
        "guestSignup": event_lookup.get("SignUp_Success_Guest", {}).get("count", 0),
        "guestSignupUsers": event_lookup.get("SignUp_Success_Guest", {}).get("users", 0),
        "fullRegistration": event_lookup.get("SignUp_Success_Registration", {}).get("count", 0),
        "fullRegistrationUsers": event_lookup.get("SignUp_Success_Registration", {}).get("users", 0),
        "taskStatusChange": event_lookup.get("task_manual_status_change", {}).get("count", 0),
        "taskStatusChangeUsers": event_lookup.get("task_manual_status_change", {}).get("users", 0),
        "taskTabClicked": event_lookup.get("task_tab_clicked", {}).get("count", 0),
        "taskTabClickedUsers": event_lookup.get("task_tab_clicked", {}).get("users", 0),
        "phaseChange": event_lookup.get("navigator_phase_change", {}).get("count", 0),
        "phaseChangeUsers": event_lookup.get("navigator_phase_change", {}).get("users", 0),
        "clickToNavigator": event_lookup.get("click_to_navigator", {}).get("count", 0),
        "clickToNavigatorUsers": event_lookup.get("click_to_navigator", {}).get("users", 0),
        "ctaClicks": event_lookup.get("call_to_action_clicks", {}).get("count", 0),
        "ctaClicksUsers": event_lookup.get("call_to_action_clicks", {}).get("users", 0),
        "outboundClicks": event_lookup.get("outbound_link_clicks", {}).get("count", 0),
        "outboundClicksUsers": event_lookup.get("outbound_link_clicks", {}).get("users", 0),
        "formSubmits": event_lookup.get("form_submits", {}).get("count", 0),
        "formSubmitsUsers": event_lookup.get("form_submits", {}).get("users", 0),
    }

    # --- Traffic Sources ---
    traffic_rows = parse_ga4_csv(ga4_traffic_file)
    traffic_summary = []
    for r in traffic_rows:
        channel = r.get("Session primary channel group (Default Channel Group)", "")
        if not channel: continue
        traffic_summary.append({
            "channel": channel,
            "sessions": safe_int(r.get("Sessions")),
            "engagedSessions": safe_int(r.get("Engaged sessions")),
            "engagementRate": round(safe_float(r.get("Engagement rate")) * 100, 1),
            "avgEngTime": round(safe_float(r.get("Average engagement time per session")), 1),
            "users": safe_int(r.get("Active users")),
            "keyEvents": safe_int(r.get("Key events")),
        })

    # --- Landing Pages (channel-level, as exported) ---
    landing_rows = parse_ga4_csv(ga4_landing_file)
    landing_summary = []
    for r in landing_rows:
        channel = r.get("First user primary channel group (Default Channel Group)", "")
        if not channel: continue
        landing_summary.append({
            "channel": channel,
            "users": safe_int(r.get("Total users")),
            "newUsers": safe_int(r.get("New users")),
            "returningUsers": safe_int(r.get("Returning users")),
            "avgEngTime": round(safe_float(r.get("Average engagement time per active user")), 1),
            "keyEvents": safe_int(r.get("Key events")),
        })

    # --- Build name → all possible task IDs from markdown files ---
    name_to_all_ids = {}
    for tcd in task_content_dirs:
        if not os.path.isdir(tcd): continue
        for f in _glob.glob(os.path.join(tcd, "*.md")):
            content = open(f).read()
            parts = content.split("---")
            if len(parts) < 3: continue
            fm = parts[1]
            task_id = task_name = None
            for line in fm.split("\n"):
                line = line.strip()
                if line.startswith("id:"):
                    task_id = line[3:].strip().strip('"').strip("'")
                elif line.startswith("name:"):
                    task_name = line[5:].strip().strip('"').strip("'")
            if task_id and task_name:
                if task_name not in name_to_all_ids:
                    name_to_all_ids[task_name] = set()
                name_to_all_ids[task_name].add(task_id)

    # --- Enrich task progress with page views ---
    id_to_url_local = {v: k for k, v in url_to_id.items()}
    for tp_row in task_progress:
        tn = tp_row["task"]
        sl = tp_row.get("slug") or task_name_to_slug.get(tn, tn)
        
        # Collect all candidate IDs for this task name
        candidates = {sl}
        candidates.add(tn)  # raw name
        for tid in name_to_all_ids.get(tn, set()):
            candidates.add(tid)
        # Also try the urlSlug if we have it from codebase
        if tp_row.get("urlSlug"):
            candidates.add(tp_row["urlSlug"])
        
        # Find the best match: the candidate with the most page views
        best_pv = None
        best_views = 0
        for cand in candidates:
            pv = task_page_views.get(cand)
            if pv and pv["views"] > best_views:
                best_pv = pv
                best_views = pv["views"]
        
        if best_pv:
            tp_row["pageViews"] = best_pv["views"]
            tp_row["pageUsers"] = best_pv["users"]
            tp_row["avgEngTime"] = best_pv["avgEngTime"]
        else:
            tp_row["pageViews"] = 0
            tp_row["pageUsers"] = 0
            tp_row["avgEngTime"] = 0

    # --- Flag stale and orphaned tasks ---
    # Build set of all reachable task slugs
    reachable_slugs = set(task_frequency.keys()) | set(task_triggers.keys()) | set(universal_tasks)
    
    # Use the codebase_task_inventory built at startup
    existing_task_ids = set(codebase_task_inventory.keys()) if codebase_task_inventory else set()

    api_task_set = set(api_task_slugs) if plurmit_data else set()
    api_aa_set = set(api_aa_slugs) if plurmit_data else set()

    stale_count = 0
    orphan_count = 0
    no_analytics_count = 0
    for tp_row in task_progress:
        tn = tp_row["task"]
        sl = tp_row.get("slug") or task_name_to_slug.get(tn, tn)
        
        # Check all possible slugs via name_to_all_ids
        all_ids = {sl, tn}
        for tid in name_to_all_ids.get(tn, set()):
            all_ids.add(tid)
        
        is_reachable = any(s in reachable_slugs for s in all_ids)
        in_codebase = any(s in existing_task_ids for s in all_ids)
        in_api = any(s in api_task_set or s in api_aa_set for s in all_ids)
        has_analytics = tp_row.get("hasAnalytics", True)
        
        # Stale: completed > page views (impossible if task page is current)
        pv = tp_row.get("pageViews", 0)
        comp = tp_row.get("completed", 0)
        is_stale = comp > pv and comp > 10
        
        # Categorize
        if is_reachable and has_analytics:
            tp_row["dataQuality"] = "ok"
        elif is_reachable and not has_analytics:
            tp_row["dataQuality"] = "noAnalytics"  # in codebase + reachable, but no XLSX data
            no_analytics_count += 1
        elif in_api:
            tp_row["dataQuality"] = "api_flow"
        elif not in_codebase:
            tp_row["dataQuality"] = "retired"
        elif in_codebase and not is_reachable:
            tp_row["dataQuality"] = "orphaned"
        else:
            tp_row["dataQuality"] = "ok"

        tp_row["stale"] = is_stale
        if is_stale: stale_count += 1
        if tp_row["dataQuality"] in ("retired", "orphaned"): orphan_count += 1

    if stale_count:
        print(f"  Likely stale tasks: {stale_count} (completed > page views)")
    if orphan_count:
        print(f"  Orphaned/retired tasks: {orphan_count} (in XLSX but not in any current roadmap)")
    if no_analytics_count:
        print(f"  Reachable but no XLSX data: {no_analytics_count} tasks (codebase-only)")

    # --- Enrich anytime actions with page views ---
    for aa in aa_reach:
        aa_id = aa["id"]
        pv = aa_page_views.get(aa_id)
        if pv:
            aa["pageViews"] = pv["views"]
            aa["pageUsers"] = pv["users"]
        else:
            aa["pageViews"] = 0
            aa["pageUsers"] = 0

    # --- Build GA4 output ---
    ga4_data = {
        "dateRange": ga4_date_range,
        "taskPageViews": task_page_views,
        "aaPageViews": aa_page_views,
        "topPages": top_pages,
        "pageCategories": page_cat_list,
        "events": events_summary,
        "funnel": funnel,
        "traffic": traffic_summary,
        "landing": landing_summary,
    }

    matched_tasks = sum(1 for sl in task_page_views if sl in task_name_to_slug.values() or sl in set(s for s in task_page_views))
    print(f"  Date range: {ga4_date_range.get('start','?')} to {ga4_date_range.get('end','?')}")
    print(f"  Page paths: {len(page_rows)}")
    print(f"  Task pages matched: {len(task_page_views)} slugs")
    print(f"  AA pages matched: {len(aa_page_views)} slugs")
    print(f"  Events: {len(events_summary)}")
    print(f"  Traffic channels: {len(traffic_summary)}")
    total_task_views = sum(v["views"] for v in task_page_views.values())
    total_aa_views = sum(v["views"] for v in aa_page_views.values())
    print(f"  Total task page views: {total_task_views:,}")
    print(f"  Total AA page views: {total_aa_views:,}")

# ============================================================
# Output
# ============================================================
output = {
    "meta": {
        "xlsxFile": os.path.basename(XLSX),
        "totalBusinesses": total_businesses,
        "totalUsers": sum(phases.values()),
    },
    "industries": combined_industries,
    "sectors": sector_summary,
    "anytimeActions": aa_reach,
    "phases": phase_summary,
    "taskProgress": task_progress,
    "nonEssentialQuestions": neq_data,
    "codebaseNEQs": codebase_neqs if nav else [],
    "legalStructures": legal_data,
    "orphanedSectors": sorted(orphaned_sectors),
    "sectorMismatches": nav_mismatches,
    "taskFrequency": task_frequency,
    "taskNameToSlug": task_name_to_slug,
    "universalTasks": universal_tasks,
    "totalDiffTasks": len(task_to_industries),
    "totalNeqTaskSlugs": len(set(t for n in codebase_neqs for t in n.get("allTasks", []))),
    "totalAllDiffTasks": len(set(list(task_to_industries.keys()) + [t for n in codebase_neqs for t in n.get("allTasks", [])])),
    "fundingCount": codebase_funding_count or len(nav.get("fundings", [])),
    "certificationCount": codebase_cert_count,
    "totalDistinctPaths": total_distinct_paths,
    "totalDistinctPathsNote": "Counts industry × legal structure variations only. NEQ combinations (56 binary questions) multiply this further.",
    "taskTriggers": task_triggers,
    "permitCoverage": plurmit_data,
    "ga4": ga4_data,
    "unknowns": {
        "industry": unknown_industry,
        "sector": unknown_sector,
        "phase": unknown_phase,
        "legalStructure": unknown_legal,
        "unmatchedIndustries": unmatched_xlsx,
        "totalAccountedFor": total_businesses,
        "totalWithUnknowns": total_businesses + unknown_phase,
    },
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w") as fh:
    json.dump(output, fh)

# Pretty version for debugging
with open(OUT.replace(".json",".pretty.json"), "w") as fh:
    json.dump(output, fh, indent=2)

print(f"\nWritten to {OUT}")
print(f"  Industries: {len(combined_industries)}")
print(f"  Sectors: {len(sector_summary)} ({len(orphaned_sectors)} orphaned)")
print(f"  Anytime Actions: {len(aa_reach)}")
print(f"  Fundings in nav: {len(nav_fundings)}")
print(f"  Phases: {len(phases)}")
print(f"  Mismatches: {len(nav_mismatches)}")
